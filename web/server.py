import argparse
import csv
import io
import json
import math
import os
from http.server import SimpleHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import urlparse

import numpy as np
from openpyxl import load_workbook


WEB_DIR = Path(__file__).resolve().parent
MAX_ROWS = 10000


def respond(handler, status, payload):
    body = json.dumps(payload, ensure_ascii=False, allow_nan=False, default=str).encode("utf-8")
    handler.send_response(status)
    handler.send_header("Content-Type", "application/json; charset=utf-8")
    handler.send_header("Content-Length", str(len(body)))
    handler.end_headers()
    handler.wfile.write(body)


def default_host():
    """Bind locally by default, but bind publicly when a platform supplies PORT."""
    return os.environ.get("HOST") or ("0.0.0.0" if os.environ.get("PORT") else "127.0.0.1")


def default_port():
    try:
        return int(os.environ.get("PORT", "8793"))
    except ValueError:
        return 8793


def finite(value):
    try:
        number = float(str(value).replace(",", "").strip())
        return number if math.isfinite(number) else None
    except (TypeError, ValueError):
        return None


def unique_headers(values):
    seen, result = {}, []
    for index, value in enumerate(values):
        base = str(value).strip() if value is not None and str(value).strip() else f"Column_{index + 1}"
        seen[base] = seen.get(base, 0) + 1
        result.append(base if seen[base] == 1 else f"{base}_{seen[base]}")
    return result


def table_payload(rows, name):
    rows = list(rows)
    if not rows:
        raise ValueError("The uploaded file is empty.")
    headers = unique_headers(rows[0])
    data = []
    for raw in rows[1 : MAX_ROWS + 1]:
        if not any(value is not None and str(value).strip() for value in raw):
            continue
        data.append({header: (raw[i] if i < len(raw) else None) for i, header in enumerate(headers)})
    if not data:
        raise ValueError("The file has headers but no data rows.")
    numeric = []
    for header in headers:
        count = sum(finite(row.get(header)) is not None for row in data)
        if count >= max(3, int(len(data) * 0.55)):
            numeric.append(header)
    return {"ok": True, "name": name, "columns": headers, "numeric_columns": numeric, "rows": data, "row_count": len(data)}


def parse_upload(raw, filename):
    suffix = Path(filename).suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        book = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        sheet = book.active
        payload = table_payload(sheet.iter_rows(values_only=True), filename)
        payload["sheet"] = sheet.title
        return payload
    if suffix == ".xls":
        raise ValueError("Legacy .xls is not supported. Save it as .xlsx or CSV first.")
    text = raw.decode("utf-8-sig", errors="replace")
    sample = text[:8192]
    try:
        dialect = csv.Sniffer().sniff(sample)
    except csv.Error:
        dialect = csv.excel
    return table_payload(csv.reader(io.StringIO(text), dialect), filename)


def ses(train, horizon):
    best = None
    for alpha in np.linspace(0.05, 0.95, 19):
        level, error = train[0], 0.0
        for value in train[1:]:
            error += (value - level) ** 2
            level = alpha * value + (1 - alpha) * level
        if best is None or error < best[0]:
            best = (error, level)
    return [float(best[1])] * horizon


def holt(train, horizon, damped=False):
    best = None
    initial_trend = float(np.mean(np.diff(train[: min(len(train), 8)])))
    for alpha in (0.2, 0.4, 0.6, 0.8):
        for beta in (0.05, 0.15, 0.3, 0.5):
            level, trend, error = train[0], initial_trend, 0.0
            for value in train[1:]:
                fitted = level + trend
                error += (value - fitted) ** 2
                old = level
                level = alpha * value + (1 - alpha) * fitted
                trend = beta * (level - old) + (1 - beta) * trend
            if best is None or error < best[0]:
                best = (error, level, trend)
    phi = 0.92 if damped else 1.0
    return [float(best[1] + best[2] * sum(phi ** j for j in range(1, i + 2))) for i in range(horizon)]


def seasonal_forecast(train, horizon, period, multiplicative=False, damped=False):
    if period < 2 or len(train) < period * 2:
        return holt(train, horizon, damped)
    trend = np.convolve(train, np.ones(period) / period, mode="valid")
    offset = period // 2
    season = np.zeros(period)
    counts = np.zeros(period)
    for i, value in enumerate(train):
        ti = min(max(i - offset, 0), len(trend) - 1)
        baseline = trend[ti]
        effect = value / baseline if multiplicative and abs(baseline) > 1e-9 else value - baseline
        season[i % period] += effect
        counts[i % period] += 1
    season = season / np.maximum(counts, 1)
    base = holt(train, horizon, damped)
    result = []
    for i, value in enumerate(base):
        effect = season[(len(train) + i) % period]
        result.append(float(value * effect if multiplicative else value + effect))
    return result


def arima_like(train, horizon, p):
    diff = np.diff(train)
    p = max(1, min(int(p), len(diff) // 3, 18))
    if len(diff) <= p + 2:
        return holt(train, horizon)
    x, y = [], []
    for i in range(p, len(diff)):
        x.append([1.0, *diff[i - p : i][::-1]])
        y.append(diff[i])
    coef = np.linalg.lstsq(np.asarray(x), np.asarray(y), rcond=None)[0]
    history, level, output = list(diff), float(train[-1]), []
    for _ in range(horizon):
        step = float(np.dot(coef, [1.0, *history[-p:][::-1]]))
        level += step
        history.append(step)
        output.append(level)
    return output


def nnetar(train, horizon, period, params):
    p = max(1, min(int(params.get("p", 6)), 24))
    seasonal_order = max(0, min(int(params.get("P", 1)), 4))
    size = max(2, min(int(params.get("size", 8)), 64))
    repeats = max(1, min(int(params.get("repeats", 8)), 30))
    ridge = max(1e-8, float(params.get("lambda", 0.01)))
    lags = list(range(1, p + 1))
    lags += [period * i for i in range(1, seasonal_order + 1) if period * i not in lags]
    max_lag = max(lags)
    if len(train) < max_lag + 8:
        return holt(train, horizon, damped=True)
    mean, scale = float(np.mean(train)), float(np.std(train) or 1.0)
    norm = [(float(v) - mean) / scale for v in train]
    x, y = [], []
    for i in range(max_lag, len(norm)):
        x.append([norm[i - lag] for lag in lags])
        y.append(norm[i])
    x, y = np.asarray(x), np.asarray(y)
    forecasts = []
    for repeat in range(repeats):
        rng = np.random.default_rng(7300 + repeat)
        weights = rng.normal(0, 0.7, (len(lags), size))
        bias = rng.normal(0, 0.25, size)
        hidden = np.tanh(x @ weights + bias)
        design = np.column_stack([np.ones(len(hidden)), hidden])
        penalty = np.eye(design.shape[1]) * ridge
        penalty[0, 0] = 0
        out = np.linalg.solve(design.T @ design + penalty, design.T @ y)
        history, path = list(norm), []
        for _ in range(horizon):
            features = np.asarray([history[-lag] for lag in lags])
            value = float(np.dot(np.r_[1.0, np.tanh(features @ weights + bias)], out))
            value = float(np.clip(value, -6, 6))
            history.append(value)
            path.append(value * scale + mean)
        forecasts.append(path)
    return np.mean(forecasts, axis=0).astype(float).tolist()


def metrics(actual, predicted):
    a, p = np.asarray(actual), np.asarray(predicted)
    rmse = float(np.sqrt(np.mean((a - p) ** 2)))
    corr = float(np.corrcoef(a, p)[0, 1]) if len(a) > 1 and np.std(a) > 0 and np.std(p) > 0 else 0.0
    if not math.isfinite(corr):
        corr = 0.0
    return {"rmse": rmse, "correlation": corr}


def run_forecast(payload):
    series = [finite(value) for value in payload.get("series", [])]
    series = [value for value in series if value is not None]
    if len(series) < 16:
        raise ValueError("Select a numeric series with at least 16 usable observations.")
    train_pct = min(90, max(55, float(payload.get("train_pct", 80))))
    split = max(8, min(len(series) - 3, round(len(series) * train_pct / 100)))
    train, actual = np.asarray(series[:split], dtype=float), series[split:]
    horizon = len(actual)
    period = max(1, min(int(payload.get("frequency", 12)), max(1, len(train) // 2)))
    selected = payload.get("models") or []
    params = payload.get("params") or {}
    ets = params.get("ets") or {}
    model_fns = {
        "SES": lambda: ses(train, horizon),
        "Holt": lambda: holt(train, horizon),
        "Holt-Winters": lambda: seasonal_forecast(train, horizon, period),
        "ARIMA": lambda: arima_like(train, horizon, params.get("nnetar", {}).get("p", 6)),
        "ETS": lambda: seasonal_forecast(
            train,
            horizon,
            period,
            multiplicative=ets.get("season") == "Multiplicative",
            damped=False,
        ),
        "NNETAR": lambda: nnetar(train, horizon, period, params.get("nnetar", {})),
    }
    results = []
    for name in selected:
        if name not in model_fns:
            continue
        prediction = model_fns[name]()
        results.append({"name": name, "forecast": prediction, **metrics(actual, prediction)})
    if not results:
        raise ValueError("Select at least one forecasting model.")
    trio = {item["name"]: item["forecast"] for item in results if item["name"] in {"ARIMA", "ETS", "NNETAR"}}
    ensemble = None
    if len(trio) == 3:
        ensemble = np.mean(np.asarray(list(trio.values())), axis=0).astype(float).tolist()
    return {
        "ok": True,
        "split": split,
        "train": train.astype(float).tolist(),
        "actual": actual,
        "results": results,
        "ensemble": ensemble,
        "ensemble_metrics": metrics(actual, ensemble) if ensemble else None,
    }


class AF3Handler(SimpleHTTPRequestHandler):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, directory=str(WEB_DIR), **kwargs)

    def log_message(self, fmt, *args):
        return

    def end_headers(self):
        self.send_header("Cache-Control", "no-store")
        super().end_headers()

    def do_GET(self):
        path = urlparse(self.path).path
        if path in {"/api/health", "/healthz"}:
            respond(self, 200, {"ok": True, "agent": "AgentFastFuriosForecaster", "codename": "AF3"})
            return
        if path == "/favicon.ico":
            self.send_response(204)
            self.end_headers()
            return
        super().do_GET()

    def do_POST(self):
        path = urlparse(self.path).path
        try:
            length = int(self.headers.get("Content-Length", "0"))
            if length > 25 * 1024 * 1024:
                raise ValueError("File is larger than the 25 MB local upload limit.")
            raw = self.rfile.read(length)
            if path == "/api/upload":
                filename = self.headers.get("X-Filename", "upload.csv")
                respond(self, 200, parse_upload(raw, filename))
            elif path == "/api/forecast":
                respond(self, 200, run_forecast(json.loads(raw.decode("utf-8"))))
            else:
                respond(self, 404, {"ok": False, "error": "Unknown API route."})
        except Exception as error:
            respond(self, 400, {"ok": False, "error": str(error)})


def main():
    parser = argparse.ArgumentParser(description="AF3 local web agent")
    parser.add_argument("--host", default=default_host())
    parser.add_argument("--port", type=int, default=default_port())
    args = parser.parse_args()
    server = ThreadingHTTPServer((args.host, args.port), AF3Handler)
    print(f"AF3 mission control: http://{args.host}:{args.port}/", flush=True)
    server.serve_forever()


if __name__ == "__main__":
    main()
